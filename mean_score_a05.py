# -*- coding: utf-8 -*-
"""
Created on Tue Jun 17 15:12:46 2023

@author: Fabrice
"""

#conda install torchvision
#import os
import torch
#import torch.nn as nn
#import torch.optim as optim
import torchvision.transforms as transforms
from torch.utils.data import DataLoader
from torchvision.datasets import ImageFolder
#from torchvision.utils import save_image
import numpy as np
#from torchvision.models import inception_v3
#from scipy.linalg import sqrtm
#from torch.autograd import Variable
#from torch.nn.utils import spectral_norm
#import torch.nn.functional as F
from training import networks_stylegan2
import openpyxl
from sklearn.metrics.pairwise import pairwise_distances as dm

image_size = 256
# transform validation data of naevus to tensor
validation_NV = ImageFolder(root="dataset/test_set/NEV",
                      transform=transforms.Compose([
                          transforms.Resize(image_size),
                          transforms.CenterCrop(image_size),
                          transforms.ToTensor(),
                          transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
                      ]))

''' 
Calculate the anomaly score  of naevus and keep it in a list
'''
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# chargement des poids du générateur
import pickle
with open('training-runs/00006-stylegan2-trainset-gpus1-batch16-gamma0.8192/network-snapshot-000560.pkl', 'rb') as f:
    generator = pickle.load(f)['G'].to(device)
generator.eval()
print("generator (G, not G_ema) weights imported")

# chargement des poids du discriminateur
with open('training-runs/00006-stylegan2-trainset-gpus1-batch16-gamma0.8192/network-snapshot-000560.pkl', 'rb') as f:
    discriminator = pickle.load(f)['D'].to(device)
discriminator.eval()
#print("Discriminator weights imported")

#import encoder weights
encoder=networks_stylegan2.encoder.to(device)
encoder.load_state_dict(torch.load("checkpointed/encoder_end.pth", map_location=torch.device('cpu')))
encoder.eval()
print("encoder weights imported")
netE=encoder

alpha = 0.5 # facteur de poids du score résiduel
k = 1 #facteur de poids du score discriminateur
NV_score = []
valNV_loader = DataLoader(validation_NV, 16, shuffle=False)
with torch.no_grad():
    for (x, label) in valNV_loader:
        bs=x.size(0)
        x=x.to(device)
        code=netE(x)
        rec_image = generator(code,0).to("cpu")
        x=x.to("cpu")
        #you can replace "manhattan" by another value, according to the type of norm
        #you want to apply: "cosine", "euclidean", "L1", "L2", etc.
        rec_diff = dm(x.view(bs, -1), rec_image.view(bs, -1), metric="manhattan")
        rec_score = np.mean(rec_diff, 1) #fait la moyenne des valeurs d'un tableau numpy sur la 2eme dim
        d_input = torch.cat((x, rec_image), 0).to(device)
        f_x, f_gx = discriminator.extract_feature(d_input,0).chunk(2,0)
        f_x = f_x.to("cpu")
        f_gx = f_gx.to("cpu")
        feat_diff = dm(f_x.view(bs, -1), f_gx.view(bs, -1), metric="manhattan")
        feat_score = np.mean(feat_diff, 1) #fait la moyenne des valeurs du tableau numpy sur la 2eme dim        
        outlier_score = alpha * rec_score + k * feat_score #score global d'anomalie
        NV_score.append(outlier_score)
        print("loops in progress for NEV_mean_score")
        #break
    NV_score = np.concatenate(NV_score)
print(len(NV_score))
#--------------------------------------------------------------------------------------------------

  # transform validation data of melanoma to tensor
validation_MEL = ImageFolder(root="dataset/test_set/MEL",
                      transform=transforms.Compose([
                          transforms.Resize(image_size),
                          transforms.CenterCrop(image_size),
                          transforms.ToTensor(),
                          transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
                      ]))

''' 
  Calculate the anomaly score of each picture of Melanomas and keep it in a list
'''

MEL_score = []
valMEL_loader = DataLoader(validation_MEL, 16, shuffle=False)

with torch.no_grad():
    for (x2, label2) in valMEL_loader:
        bs2=x2.size(0)
        x2=x2.to(device)
        code2=netE(x2)
        rec2_image = generator(code2,0).to("cpu")
        x2=x2.to("cpu")
        rec2_diff = dm(x2.view(bs2, -1), rec2_image.view(bs2, -1), metric="manhattan")
        #you can replace "manhattan" by another value, according to the type of norm
        #you want to apply: "cosine", "euclidean", "L1", "L2", etc.
        rec2_score = np.mean(rec2_diff, 1) #fait la moyenne des valeurs d'un tableau numpy sur la 2eme dim
        d_input2 = torch.cat((x2, rec2_image), 0).to(device)
        f_x2, f_gx2 = discriminator.extract_feature(d_input2,0).chunk(2,0)
        f_x2 = f_x2.to("cpu")
        f_gx2 = f_gx2.to("cpu")
        feat2_diff = dm(f_x2.view(bs2, -1), f_gx2.view(bs2, -1), metric="manhattan")
        feat2_score = np.mean(feat2_diff, 1) #fait la moyenne des valeurs du tableau numpy sur la 2eme dim        
        outlier2_score = (alpha * rec2_score + k * feat2_score) #score global d'anomalie
        MEL_score.append(outlier2_score)
        print("loops in progress for MEL_mean_score")
        #break
    MEL_score = np.concatenate(MEL_score)
print(len(MEL_score))
''' 
Save both lists in an excel file in which naevus label is 0 and melanoma label is 1
'''

def excel_file(y_score1, y_score2):
    # Crate a new excel file
    classeur = openpyxl.Workbook()
    feuille = classeur.active

    # Add a title for scores and labels
    feuille.cell(row=1, column=1, value="score")
    feuille.cell(row=1, column=2, value="label")

    # Add elements of y_score1 in column 1 and its label '0' in column 2
    for i, score in enumerate(y_score1):
        feuille.cell(row=i+2, column=1, value=score)
        feuille.cell(row=i+2, column=2, value=0)
      # Add elements of y_score2 "after y_score1" in column 1 and its label '0' in column 2
    for i, score in enumerate(y_score2):
        feuille.cell(row=i+len(y_score1)+2, column=1, value=score)
        feuille.cell(row=i+len(y_score1)+2, column=2, value=1)
      # save file
    classeur.save("score_a05_manhattan.xlsx")

# Save NV_score and MEL_score in excel file
y_score1 = NV_score
y_score2 = MEL_score

excel_file(y_score1, y_score2)
