# -*- coding: utf-8 -*-
"""
Created on Fri Aug 18 10:07:17 2023

@author: Fabrice
"""

import os
import torch
import pandas as pd
import numpy as np
from training import networks_stylegan2
import torchvision.transforms as transforms
from torch.utils.data import DataLoader
from torchvision.datasets import ImageFolder
import pickle
import openpyxl
import torch.nn as nn

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
batch_size = 16
image_size = 256
    

#load generator weight
with open('training-runs/00006-stylegan2-trainset-gpus1-batch16-gamma0.8192/network-snapshot-000560.pkl', 'rb') as f:
    generator = pickle.load(f)['G'].to(device)
generator.eval()

#load discriminator weight
with open('training-runs/00006-stylegan2-trainset-gpus1-batch16-gamma0.8192/network-snapshot-000560.pkl', 'rb') as f:
    discriminator = pickle.load(f)['D'].to(device)
discriminator.eval()

# load encoder weights
encoder=networks_stylegan2.encoder.to(device)
encoder.load_state_dict(torch.load("checkpointed/encoder_end.pth", map_location=torch.device('cpu')))
encoder.eval()


def averageMeanScore(input_dir):
    datatensor = ImageFolder(input_dir,
                          transform=transforms.Compose([
                              transforms.Resize(image_size),
                              transforms.CenterCrop(image_size),
                              transforms.ToTensor(),
                              transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
                          ]))
    
    dataloader = DataLoader(datatensor, batch_size=batch_size, shuffle=False)
    
    # fetch image name's in subdirectories of input_dir
    list_name = [os.path.basename(path) for path, _ in datatensor.imgs]
    
    #initialise empty lists for latent vectors
    latent_vectors = []
    
    #construct a list of all latent vectors
    with torch.no_grad():
        for (x, _) in dataloader:
            x=x.to(device)
            code=encoder(x) 
            for vec in code:           
                latent_vectors.append(vec.cpu().numpy())
        
    #associate each latent vector with the name of image it comes from, in a dataframe    
    latent_df = pd.DataFrame(latent_vectors, columns=[f"dim_{i+1}" for i in range(512)])
    latent_df['image_name'] = list_name
       
    # initialize a List to store all average distances
    all_average_distances = []
    
    # Browse both datatensor and latent_df
    for image_index, (image_start, _) in enumerate(datatensor):
        image_start = image_start.unsqueeze(0).to(device)
    
        # Latent vector corresponding to the starting image, retrieved in latent_df
        latent_vector_start = latent_df.iloc[image_index, :-1].values.astype(np.float32)
        latent_tensor_start = torch.tensor(latent_vector_start, dtype=torch.float32).to(device)
        
        # Calculer les distances entre le vecteur de départ et tous les autres vecteurs latents
        with torch.no_grad():
            distances = []
            pwd = nn.PairwiseDistance(p=2, keepdim=True)
            for _, row in latent_df.iterrows():
                latent_vector = row.iloc[:-1].values.astype(np.float32)
                latent_tensor = torch.tensor(latent_vector, dtype=torch.float32).to(device)
                distance_lt = pwd(latent_tensor_start, latent_tensor).to("cpu")
                distances.append(distance_lt.item())
    
            # Sort image clues based on distances
            sorted_indices = np.argsort(distances)
    
            # Choose the 5 closest vectors (excluding the starting vector)
            nearest_indices = sorted_indices[1:6]
            
            # Liste pour stocker les distances entre vec(x) et ses voisins
            knn_distances = []
            # Générer les images à partir des vecteurs latents les plus proches
            for index in nearest_indices:
                nearest_latent_vector = latent_df.iloc[index, :-1].values.astype(np.float32)
                nearest_latent_tensor = torch.tensor(nearest_latent_vector, dtype=torch.float32).to(device)
                
                ## Generate pictures from latent vectors
                generated_image = generator(nearest_latent_tensor.unsqueeze(0), 0)
                generated_image = generated_image * 0.5 + 0.5
                
                
                #Calculate the distance between image features (x) 
                #and its closest reconstructed neighbors
                d_input = torch.cat((image_start, generated_image), 0)
                f_x, f_gx = discriminator.extract_feature(d_input,0).chunk(2,0)
                distance_latent = pwd(f_x.view(-1), f_gx.view(-1)).to("cpu")
                knn_distances.append(distance_latent.item())
            
            # Calculate the average of the distances between the starting image and the generated images
            average_distance = np.mean(knn_distances)
            print(list_name[image_index],"=",average_distance)
            # Add mean distances to the global distances list
            all_average_distances.append(average_distance)
    return(all_average_distances)
            

avgScore_NEV = averageMeanScore(input_dir="dataset/test_set/NEV")
avgScore_MEL = averageMeanScore(input_dir="dataset/test_set/MEL")
def excel_file(y_score1, y_score2):
    # Crate a new excel file
    classeur = openpyxl.Workbook()
    feuille = classeur.active

    # Add a title for scores and labels
    feuille.cell(row=1, column=1, value="score")
    feuille.cell(row=1, column=2, value="label")

    # Add elements of y_score1 in column 1 and its label '0' in column 2
    for i, score in enumerate(y_score1):
        feuille.cell(row=i+2, column=1, value=score)
        feuille.cell(row=i+2, column=2, value=0)

    # Add elements of y_score2 "after y_score1" in column 1 and its label '0' in column 2
    for i, score in enumerate(y_score2):
        feuille.cell(row=i+len(y_score1)+2, column=1, value=score)
        feuille.cell(row=i+len(y_score1)+2, column=2, value=1)

    # save file
    classeur.save("avgFeatScore_ntwk560.xlsx")

# Save NV_score and MEL_score in excel file
y_score1 = avgScore_NEV
y_score2 = avgScore_MEL

excel_file(y_score1, y_score2)